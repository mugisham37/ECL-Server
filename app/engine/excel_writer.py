"""Excel output workbook generation for ECL compute runs."""

from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from app.engine.pd_engine import STAGES

METADATA_LABELS = (
    "Run ID",
    "Tenant ID",
    "Timestamp UTC",
    "Engine Version",
    "Input Hashes",
)

_CURRENCY_FMT = "#,##0.00"
_PERCENT_FMT = "0.00%"
_PROB_FMT = "0.00000000"
_HEADER_FONT = Font(bold=True)


def _workbook_to_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_metadata(ws, run_meta: dict[str, Any]) -> None:
    values = (
        run_meta["run_id"],
        run_meta["tenant_id"],
        run_meta["timestamp_utc"],
        run_meta["engine_version"],
        run_meta["input_hashes"],
    )
    for col_idx, (label, value) in enumerate(zip(METADATA_LABELS, values, strict=True), start=1):
        ws.cell(row=1, column=col_idx, value=label).font = _HEADER_FONT
        ws.cell(row=2, column=col_idx, value=value)


def _write_dataframe(
    ws,
    df: pd.DataFrame,
    *,
    start_row: int,
    currency_columns: set[str] | None = None,
    percent_columns: set[str] | None = None,
    prob_columns: set[str] | None = None,
) -> None:
    currency_columns = currency_columns or set()
    percent_columns = percent_columns or set()
    prob_columns = prob_columns or set()

    for row_idx, row in enumerate(
        dataframe_to_rows(df, index=False, header=True),
        start=start_row,
    ):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == start_row:
                cell.font = _HEADER_FONT
                continue
            if df.columns[col_idx - 1] in currency_columns:
                cell.number_format = _CURRENCY_FMT
            elif df.columns[col_idx - 1] in percent_columns:
                cell.number_format = _PERCENT_FMT
            elif df.columns[col_idx - 1] in prob_columns:
                cell.number_format = _PROB_FMT

    for col_idx in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18


def _matrix_rows(
    matrices: dict[str, dict[str, list[float]]],
    *,
    include_month: bool = False,
    monthly: dict[str, dict[str, dict[str, list[float]]]] | None = None,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    if include_month and monthly is not None:
        for segment in sorted(monthly.keys()):
            for month_key in sorted(monthly[segment].keys()):
                matrix = monthly[segment][month_key]
                for from_stage in STAGES:
                    values = matrix[from_stage]
                    rows.append(
                        {
                            "Segment": segment,
                            "Month": month_key,
                            "From Stage": from_stage,
                            "To Stage 1": values[0],
                            "To Stage 2": values[1],
                            "To Stage 3": values[2],
                        }
                    )
        return pd.DataFrame(rows)

    for segment in sorted(matrices.keys()):
        matrix = matrices[segment]
        for from_stage in STAGES:
            values = matrix[from_stage]
            rows.append(
                {
                    "Segment": segment,
                    "From Stage": from_stage,
                    "To Stage 1": values[0],
                    "To Stage 2": values[1],
                    "To Stage 3": values[2],
                }
            )
    return pd.DataFrame(rows)


def _build_pd_calcs_workbook(run_meta: dict[str, Any], pd_data: dict[str, Any]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    intermediates = pd_data["intermediates"]
    pd_df = pd_data["pd_df"]

    sheet_specs: list[tuple[str, pd.DataFrame, set[str] | None]] = [
        (
            "Monthly Transition",
            _matrix_rows(
                {},
                include_month=True,
                monthly=intermediates["monthly_matrices"],
            ),
            {"To Stage 1", "To Stage 2", "To Stage 3"},
        ),
        (
            "Transition",
            _matrix_rows(intermediates["aggregate"]),
            {"To Stage 1", "To Stage 2", "To Stage 3"},
        ),
        (
            "Proportion Matrices",
            _matrix_rows(intermediates["proportion"]),
            {"To Stage 1", "To Stage 2", "To Stage 3"},
        ),
        (
            "Normalized",
            _matrix_rows(intermediates["normalized"]),
            {"To Stage 1", "To Stage 2", "To Stage 3"},
        ),
        (
            "Marginal PD",
            pd_df.rename(
                columns={
                    "SEGMENT": "Segment",
                    "Month": "Month",
                    "Transition": "Transition",
                    "Stage_1_prob": "Stage 1 Prob",
                    "Stage_2_prob": "Stage 2 Prob",
                    "Stage_3_prob": "Stage 3 Prob",
                    "Marginal_PD": "Marginal PD",
                    "Cure_Rate": "Cure Rate",
                }
            ),
            {
                "Stage 1 Prob",
                "Stage 2 Prob",
                "Stage 3 Prob",
                "Marginal PD",
                "Cure Rate",
            },
        ),
    ]

    for title, df, prob_cols in sheet_specs:
        ws = wb.create_sheet(title=title)
        _write_metadata(ws, run_meta)
        _write_dataframe(ws, df, start_row=4, prob_columns=prob_cols)

    return _workbook_to_bytes(wb)


def _split_lgd_frames(lgd_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    sum_col = "Sum of Discounted Collaterals per Loan ID"
    discounted_value_cols = [col for col in lgd_df.columns if col.startswith("Discounted ")]
    per_loan_cols = [col for col in lgd_df.columns if col not in discounted_value_cols]
    discounted_cols = [
        col for col in ("Loan ID", "Customer ID", *discounted_value_cols, sum_col) if col in lgd_df.columns
    ]
    return lgd_df[per_loan_cols].copy(), lgd_df[discounted_cols].copy()


def _build_lgd_workbook(run_meta: dict[str, Any], lgd_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    per_loan, discounted = _split_lgd_frames(lgd_df)
    currency_cols = {
        col
        for col in per_loan.columns
        if col
        not in {"Loan ID", "Customer ID", "EIR", "Proportion", "Total Loan Amount per Customer"}
    }

    for title, df, cols in (
        ("Per-Loan Collateral", per_loan, currency_cols | {"EIR", "Proportion"}),
        ("Discounted Collateral", discounted, set(discounted.columns) - {"Loan ID", "Customer ID"}),
    ):
        ws = wb.create_sheet(title=title)
        _write_metadata(ws, run_meta)
        percent_cols = {"EIR", "Proportion"} & cols
        money_cols = cols - percent_cols
        _write_dataframe(
            ws,
            df,
            start_row=4,
            currency_columns=money_cols,
            percent_columns=percent_cols,
        )

    return _workbook_to_bytes(wb)


def _build_rundown_workbook(run_meta: dict[str, Any], ead_df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rundown"
    _write_metadata(ws, run_meta)

    currency_cols = {
        "Monthly Instalment",
        "Balance After Repayment",
        "Balance After Missed Payment",
        "Sum of Discounted Collaterals",
        "Credit Loss",
        "Discounted ECL",
    }
    prob_cols = {"Marginal PD", "Cure Rate", "LGW", "LGD"}
    _write_dataframe(
        ws,
        ead_df,
        start_row=4,
        currency_columns=currency_cols,
        percent_columns=prob_cols,
    )
    return _workbook_to_bytes(wb)


def _build_ecl_summary_workbook(
    run_meta: dict[str, Any],
    aggregates: dict[str, Any],
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    ecl_by_loan = aggregates["ecl_by_loan"]
    ecl_by_segment = aggregates["ecl_by_segment"]
    ecl_by_stage = aggregates["ecl_by_stage"]
    ecl_total = aggregates["ecl_total"]

    specs: list[tuple[str, pd.DataFrame, set[str], set[str]]] = [
        (
            "ECL Per Loan",
            ecl_by_loan,
            {"Total Discounted ECL", "Total Outstanding"},
            set(),
        ),
        (
            "ECL By Segment",
            ecl_by_segment,
            {"Total ECL", "Total Outstanding"},
            {"Coverage"},
        ),
        (
            "ECL By Stage",
            ecl_by_stage,
            {"Total ECL"},
            {"Portfolio Share"},
        ),
        (
            "ECL Total",
            pd.DataFrame([ecl_total]),
            {"Total ECL", "Total Outstanding"},
            {"Coverage Ratio"},
        ),
    ]

    for title, df, currency_cols, percent_cols in specs:
        ws = wb.create_sheet(title=title)
        _write_metadata(ws, run_meta)
        _write_dataframe(
            ws,
            df,
            start_row=4,
            currency_columns=currency_cols,
            percent_columns=percent_cols,
        )

    return _workbook_to_bytes(wb)


def build_ecl_aggregates(ead_df: pd.DataFrame, ead_input_df: pd.DataFrame) -> dict[str, Any]:
    """Build aggregate DataFrames for the ECL Summary workbook and run totals."""
    if ead_df.empty:
        empty_loan = pd.DataFrame(
            columns=[
                "Loan ID",
                "Customer ID",
                "Segment",
                "Stage",
                "Total Discounted ECL",
            ]
        )
        empty_segment = pd.DataFrame(
            columns=["Segment", "Total ECL", "Total Outstanding", "Coverage"]
        )
        empty_stage = pd.DataFrame(columns=["Stage", "Total ECL", "Portfolio Share"])
        return {
            "ecl_by_loan": empty_loan,
            "ecl_by_segment": empty_segment,
            "ecl_by_stage": empty_stage,
            "ecl_total": {
                "Total ECL": 0.0,
                "Total Outstanding": 0.0,
                "Coverage Ratio": 0.0,
                "Run ID": "",
                "Engine Version": "",
                "Timestamp UTC": "",
            },
            "total_ecl": 0.0,
            "total_outstanding": 0.0,
            "coverage_ratio": 0.0,
        }

    loan_ecl = (
        ead_df.groupby(["Loan ID", "Customer ID", "SEGMENT", "Staging"], sort=True)["Discounted ECL"]
        .sum()
        .reset_index(name="Total Discounted ECL")
        .rename(columns={"SEGMENT": "Segment", "Staging": "Stage"})
    )

    outstanding_by_loan = (
        ead_input_df.drop_duplicates(subset=["Loan ID"], keep="first")
        .set_index("Loan ID")["Outstanding Amount"]
        .astype(float)
    )
    loan_ecl["Total Outstanding"] = loan_ecl["Loan ID"].map(outstanding_by_loan).fillna(0.0)

    segment_ecl = (
        loan_ecl.groupby("Segment", sort=True)["Total Discounted ECL"]
        .sum()
        .reset_index(name="Total ECL")
    )
    segment_outstanding = (
        ead_input_df.groupby("SEGMENT", sort=True)["Outstanding Amount"]
        .sum()
        .reset_index(name="Total Outstanding")
        .rename(columns={"SEGMENT": "Segment"})
    )
    ecl_by_segment = segment_ecl.merge(segment_outstanding, on="Segment", how="left")
    ecl_by_segment["Coverage"] = ecl_by_segment.apply(
        lambda row: (float(row["Total ECL"]) / float(row["Total Outstanding"]))
        if float(row["Total Outstanding"]) > 0
        else 0.0,
        axis=1,
    )

    total_ecl = float(loan_ecl["Total Discounted ECL"].sum())
    total_outstanding = float(outstanding_by_loan.sum())
    coverage_ratio = (total_ecl / total_outstanding) if total_outstanding > 0 else 0.0

    stage_ecl = (
        loan_ecl.groupby("Stage", sort=True)["Total Discounted ECL"]
        .sum()
        .reset_index(name="Total ECL")
    )
    stage_ecl["Portfolio Share"] = stage_ecl["Total ECL"].apply(float) / total_ecl if total_ecl > 0 else 0.0

    return {
        "ecl_by_loan": loan_ecl,
        "ecl_by_segment": ecl_by_segment,
        "ecl_by_stage": stage_ecl,
        "ecl_total": {
            "Total ECL": total_ecl,
            "Total Outstanding": total_outstanding,
            "Coverage Ratio": coverage_ratio,
            "Run ID": "",
            "Engine Version": "",
            "Timestamp UTC": "",
        },
        "total_ecl": total_ecl,
        "total_outstanding": total_outstanding,
        "coverage_ratio": coverage_ratio,
    }


def generate_all_workbooks(
    run_meta: dict[str, Any],
    pd_data: dict[str, Any],
    lgd_df: pd.DataFrame,
    ead_df: pd.DataFrame,
    aggregates: dict[str, Any],
) -> dict[str, bytes]:
    """Generate the four output workbooks and return filename → bytes."""
    ecl_total = aggregates["ecl_total"]
    ecl_total["Run ID"] = run_meta["run_id"]
    ecl_total["Engine Version"] = run_meta["engine_version"]
    ecl_total["Timestamp UTC"] = run_meta["timestamp_utc"]

    return {
        "PD Calcs.xlsx": _build_pd_calcs_workbook(run_meta, pd_data),
        "LGD.xlsx": _build_lgd_workbook(run_meta, lgd_df),
        "Contractual Rundown.xlsx": _build_rundown_workbook(run_meta, ead_df),
        "ECL Summary.xlsx": _build_ecl_summary_workbook(run_meta, aggregates),
    }
