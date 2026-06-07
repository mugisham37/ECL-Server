#!/usr/bin/env python3
"""Generate comprehensive IFRS 9 ECL input templates (PD, LGD, EAD).

Run from the ECL-Server root:
    python scripts/generate_templates.py
"""

from __future__ import annotations

import sys
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
_OUT_DIR = _REPO_ROOT / "app" / "static" / "templates"

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    sys.exit("openpyxl is required.  Run: pip install openpyxl")


# ── Palette / constants ────────────────────────────────────────────────────────
_HDR_BG      = "1A3C5E"   # ECL dark navy
_HDR_FG      = "FFFFFF"
_ALT_ROW     = "F0F7FF"   # light blue alternating rows
_REQ_TINT    = "D9E8F5"   # required-column tint in Instructions
_BORDER_CLR  = "94A3B8"

_TAB_PD  = "2563EB"
_TAB_LGD = "16A34A"
_TAB_EAD = "9333EA"

_FMT_DATE     = "YYYY-MM-DD"
_FMT_CURRENCY = '#,##0.00'
_FMT_DEC4     = '0.0000'

_DV_STAGE = '"Stage 1,Stage 2,Stage 3"'
_DV_FREQ  = '"MTH,QTR"'


# ── Style helpers ──────────────────────────────────────────────────────────────
def _font(bold: bool = False, size: int = 10, color: str = "000000") -> Font:
    return Font(name="Calibri", bold=bold, size=size, color=color)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _thin_border() -> Border:
    s = Side(style="thin", color=_BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_border() -> Border:
    thin = Side(style="thin",   color=_BORDER_CLR)
    thick = Side(style="medium", color=_HDR_BG)
    return Border(left=thin, right=thin, top=thin, bottom=thick)


def _align(h: str = "left", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


def _col_letter(n: int) -> str:
    return get_column_letter(n)


def _style_header_row(ws, n_cols: int) -> None:
    ws.row_dimensions[1].height = 30
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.font      = _font(bold=True, size=11, color=_HDR_FG)
        cell.fill      = _fill(_HDR_BG)
        cell.alignment = _align("center", wrap=True)
        cell.border    = _header_border()


def _style_data_rows(ws, first: int, last: int, n_cols: int) -> None:
    for r in range(first, last + 1):
        bg = _ALT_ROW if r % 2 == 0 else "FFFFFF"
        ws.row_dimensions[r].height = 18
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill   = _fill(bg)
            cell.border = _thin_border()


def _add_dv(ws, formula: str, sqref: str, title: str, prompt: str) -> None:
    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,
        showInputMessage=True,
        promptTitle=title,
        prompt=prompt,
        showErrorMessage=True,
        errorTitle="Invalid value",
        error=f"Choose from the list: {title}",
    )
    ws.add_data_validation(dv)
    dv.sqref = sqref


def _col_width(ws, col: int, width: float) -> None:
    ws.column_dimensions[_col_letter(col)].width = width


# ── Instructions sheet factory ─────────────────────────────────────────────────
def _make_instructions(
    wb: Workbook,
    tab_color: str,
    title: str,
    intro: str,
    col_rows: list[tuple],  # (name, required, type_, allowed, rule, example)
    extra: list[tuple[str, list[str]]] | None = None,
) -> None:
    ws = wb.create_sheet("Instructions")
    ws.sheet_properties.tabColor = tab_color
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 50
    ws.column_dimensions["F"].width = 28

    def _write_banner(row: int, text: str, merge: bool = True) -> None:
        ws.row_dimensions[row].height = 28
        c = ws.cell(row=row, column=1, value=text)
        c.font      = _font(bold=True, size=13, color=_HDR_FG)
        c.fill      = _fill(_HDR_BG)
        c.alignment = _align("left")
        if merge:
            ws.merge_cells(f"A{row}:F{row}")

    def _write_note(row: int, text: str) -> None:
        ws.row_dimensions[row].height = 40
        c = ws.cell(row=row, column=1, value=text)
        c.font      = _font(size=10, color="334155")
        c.fill      = _fill("EFF6FF")
        c.alignment = _align("left", wrap=True)
        c.border    = _thin_border()
        ws.merge_cells(f"A{row}:F{row}")

    _write_banner(1, title)
    _write_note(2, intro)

    # Table header row
    tbl_hdr_row = 4
    ws.row_dimensions[tbl_hdr_row].height = 22
    for ci, label in enumerate(
        ["Column Name", "Required", "Type", "Allowed Values", "Business Rule", "Example"], 1
    ):
        c = ws.cell(row=tbl_hdr_row, column=ci, value=label)
        c.font      = _font(bold=True, size=10, color=_HDR_FG)
        c.fill      = _fill(_HDR_BG)
        c.alignment = _align("center", wrap=True)
        c.border    = _header_border()

    for i, (name, req, typ, allowed, rule, example) in enumerate(col_rows):
        r = tbl_hdr_row + 1 + i
        ws.row_dimensions[r].height = 36
        vals = [name, req, typ, allowed, rule, example]
        bg = _REQ_TINT if req == "Yes" else "FAFAFA"
        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=val)
            c.fill      = _fill(bg)
            c.font      = _font(bold=(ci == 1))
            c.alignment = _align("left", wrap=True)
            c.border    = _thin_border()

    next_row = tbl_hdr_row + 1 + len(col_rows) + 1

    if extra:
        for section_title, lines in extra:
            _write_banner(next_row, section_title)
            next_row += 1
            for line in lines:
                ws.row_dimensions[next_row].height = 24
                c = ws.cell(row=next_row, column=1, value=line)
                c.font      = _font(size=10, color="1e293b")
                c.fill      = _fill("F8FAFC")
                c.alignment = _align("left", wrap=True)
                c.border    = _thin_border()
                ws.merge_cells(f"A{next_row}:F{next_row}")
                next_row += 1
            next_row += 1


# ── PD template ───────────────────────────────────────────────────────────────
def _build_pd(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Data"
    ws.sheet_properties.tabColor = _TAB_PD

    headers = ["Loan ID", "SEGMENT", "Reporting Month", "Staging", "Loan Amount"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    # 5 loans × 3 months showing staging transitions
    rows = [
        # Loan ID,         SEGMENT,      Reporting Month,  Staging,   Loan Amount
        ("LN-2025-0001", "Retail",      "2025-01-31",    "Stage 1",  500_000.00),
        ("LN-2025-0002", "Retail",      "2025-01-31",    "Stage 1",  750_000.00),
        ("LN-2025-0003", "Corporate",   "2025-01-31",    "Stage 1", 3_200_000.00),
        ("LN-2025-0004", "SME",         "2025-01-31",    "Stage 2", 1_100_000.00),
        ("LN-2025-0005", "SME",         "2025-01-31",    "Stage 1",   880_000.00),
        # February — LN-0002 deteriorates to Stage 2
        ("LN-2025-0001", "Retail",      "2025-02-28",    "Stage 1",  495_000.00),
        ("LN-2025-0002", "Retail",      "2025-02-28",    "Stage 2",  750_000.00),
        ("LN-2025-0003", "Corporate",   "2025-02-28",    "Stage 1", 3_150_000.00),
        ("LN-2025-0004", "SME",         "2025-02-28",    "Stage 2", 1_080_000.00),
        ("LN-2025-0005", "SME",         "2025-02-28",    "Stage 1",   870_000.00),
        # March — LN-0004 deteriorates to Stage 3, LN-0002 remains Stage 2
        ("LN-2025-0001", "Retail",      "2025-03-31",    "Stage 1",  490_000.00),
        ("LN-2025-0002", "Retail",      "2025-03-31",    "Stage 2",  748_000.00),
        ("LN-2025-0003", "Corporate",   "2025-03-31",    "Stage 1", 3_100_000.00),
        ("LN-2025-0004", "SME",         "2025-03-31",    "Stage 3", 1_060_000.00),
        ("LN-2025-0005", "SME",         "2025-03-31",    "Stage 1",   860_000.00),
    ]

    for r, (loan_id, segment, month, staging, amount) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=loan_id).alignment = _align("center")
        ws.cell(row=r, column=2, value=segment).alignment  = _align("center")
        d = ws.cell(row=r, column=3, value=month)
        d.number_format = _FMT_DATE
        d.alignment = _align("center")
        ws.cell(row=r, column=4, value=staging).alignment  = _align("center")
        a = ws.cell(row=r, column=5, value=amount)
        a.number_format = _FMT_CURRENCY
        a.alignment = _align("right")

    n_data = len(rows)
    _style_header_row(ws, 5)
    _style_data_rows(ws, 2, 1 + n_data, 5)
    ws.freeze_panes = "A2"

    for w, col in zip([20, 16, 20, 16, 20], range(1, 6)):
        _col_width(ws, col, w)

    _add_dv(ws, _DV_STAGE, "D2:D50000", "Staging", "Stage 1 = performing  |  Stage 2 = SICR  |  Stage 3 = default")

    _make_instructions(
        wb,
        tab_color=_TAB_PD,
        title="Probability of Default (PD) — Upload Guide",
        intro=(
            "Upload one or more monthly PD files per run. Each file is a snapshot of your loan book for "
            "one or more reporting months. The engine automatically derives month-to-month staging "
            "transitions to build an IFRS 9 3×3 transition matrix per segment.  Supply at least "
            "2–3 months of history to obtain meaningful lifetime PD curves."
        ),
        col_rows=[
            ("Loan ID",         "Yes", "Text",   "Any non-empty string",
             "Must be non-empty. Each (Loan ID, Reporting Month) pair must be unique across all "
             "uploaded PD files for this run — duplicates trigger EC-10.",
             "LN-2025-0001"),
            ("SEGMENT",         "Yes", "Text",   "Must match a configured segment name",
             "Case-sensitive match to one of your workspace's Segment names "
             "(Workspace → Admin → Segments).",
             "Retail"),
            ("Reporting Month", "Yes", "Date",   "YYYY-MM-DD or Excel date",
             "The as-at date for this row. Typically the last calendar day of the reporting month.",
             "2025-01-31"),
            ("Staging",         "Yes", "Enum",   "Stage 1  /  Stage 2  /  Stage 3",
             "IFRS 9 credit stage at the reporting date. Case-sensitive. "
             "Stage 1 = performing, Stage 2 = significant increase in credit risk (SICR), "
             "Stage 3 = credit-impaired / default.",
             "Stage 1"),
            ("Loan Amount",     "Yes", "Number", "≥ 0",
             "Outstanding principal balance at the reporting date. Must be non-negative.",
             "500000.00"),
        ],
        extra=[
            ("How the transition matrix is constructed", [
                "The engine joins each loan's staging in month N with its staging in month N+1.",
                "  • LN-001 is Stage 1 in Jan-25 and Stage 2 in Feb-25  →  Stage 1→Stage 2 transition recorded.",
                "  • LN-001 is Stage 1 in Jan-25 and absent in Feb-25  →  Stage 1→Offbooks (fully repaid/written off).",
                "All monthly transitions are summed into a 3×3 aggregate matrix per SEGMENT.",
                "Row proportions are computed, Stage 1 and Stage 3 rows are normalised (IFRS 9 boundary conditions),",
                "and the resulting matrix is raised to powers 1–299 to produce the lifetime PD curve.",
                "Cure rate = P(Stage 3 → Stage 1) + P(Stage 3 → Stage 2) from the proportion matrix.",
            ]),
            ("Accepted aliases for Reporting Month column", [
                "The engine also recognises the column name as:   Reporting Month (\"As At\")   or   As At",
            ]),
            ("Multiple files per run (PD only)", [
                "PD is the only input type where multiple files can be uploaded per run (up to 10 files, 25 MB each).",
                "All files are concatenated and deduplicated on (Loan ID, Reporting Month) before the engine runs.",
                "This allows you to split large monthly exports into separate files.",
            ]),
        ],
    )


# ── LGD template ──────────────────────────────────────────────────────────────
def _build_lgd(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Data"
    ws.sheet_properties.tabColor = _TAB_LGD

    headers = [
        "Customer ID",
        "Loan ID",
        "Outstanding Amount",
        "Effective Interest Rate (EIR)",
        "Real Estate",
        "Motor Vehicle",
        "Cash Deposit",
        "Corporate Guarantee",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    rows = [
        # CustID       LoanID           OutAmt          EIR    Real Estate  Motor Veh  Cash Dep  Corp Guar
        ("CUST-001", "LN-2025-0001",  490_000.00,  0.1400,  350_000.0,       0.0,       0.0,      0.0),
        ("CUST-001", "LN-2025-0002",  748_000.00,  0.1400,        0.0,  250_000.0,       0.0,      0.0),
        ("CUST-002", "LN-2025-0003", 3_100_000.00, 0.1200, 2_500_000.0,       0.0,       0.0, 800_000.0),
        ("CUST-003", "LN-2025-0004", 1_060_000.00, 0.1600,  600_000.0,  180_000.0,       0.0,      0.0),
        ("CUST-003", "LN-2025-0005",  860_000.00,  0.1600,        0.0,       0.0,  200_000.0,      0.0),
        ("CUST-004", "LN-2025-0006", 1_250_000.00, 0.1300,  900_000.0,       0.0,       0.0,      0.0),
        ("CUST-005", "LN-2025-0007",  425_000.00,  0.1500,        0.0,  120_000.0,       0.0,      0.0),
        ("CUST-005", "LN-2025-0008",  980_000.00,  0.1500,  750_000.0,       0.0,       0.0, 100_000.0),
        ("CUST-006", "LN-2025-0009", 2_200_000.00, 0.1100, 1_800_000.0,       0.0,  300_000.0,      0.0),
        ("CUST-007", "LN-2025-0010",  630_000.00,  0.1700,        0.0,  200_000.0,       0.0,      0.0),
    ]

    for r, (cid, lid, out, eir, re, mv, cd, cg) in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=cid).alignment  = _align("center")
        ws.cell(row=r, column=2, value=lid).alignment  = _align("center")
        out_c = ws.cell(row=r, column=3, value=out)
        out_c.number_format = _FMT_CURRENCY
        out_c.alignment = _align("right")
        eir_c = ws.cell(row=r, column=4, value=eir)
        eir_c.number_format = _FMT_DEC4
        eir_c.alignment = _align("center")
        for col, val in zip(range(5, 9), [re, mv, cd, cg]):
            c_cell = ws.cell(row=r, column=col, value=val)
            c_cell.number_format = _FMT_CURRENCY
            c_cell.alignment = _align("right")

    _style_header_row(ws, 8)
    _style_data_rows(ws, 2, 11, 8)
    ws.freeze_panes = "A2"

    for w, col in zip([16, 18, 22, 32, 16, 16, 14, 22], range(1, 9)):
        _col_width(ws, col, w)

    # Collateral Config sheet
    ws_cc = wb.create_sheet("Collateral Config")
    ws_cc.sheet_properties.tabColor = _TAB_LGD
    ws_cc.column_dimensions["A"].width = 34
    ws_cc.column_dimensions["B"].width = 58
    ws_cc.column_dimensions["C"].width = 28

    def _banner(row: int, text: str) -> None:
        ws_cc.row_dimensions[row].height = 28
        c = ws_cc.cell(row=row, column=1, value=text)
        c.font      = _font(bold=True, size=13, color=_HDR_FG)
        c.fill      = _fill(_HDR_BG)
        c.alignment = _align("left")
        ws_cc.merge_cells(f"A{row}:C{row}")

    def _row(row: int, label: str, value: str, bold_label: bool = False) -> None:
        ws_cc.row_dimensions[row].height = 36
        lc = ws_cc.cell(row=row, column=1, value=label)
        lc.font      = _font(bold=bold_label, size=10)
        lc.fill      = _fill("EFF6FF")
        lc.alignment = _align("left", wrap=True)
        lc.border    = _thin_border()
        vc = ws_cc.cell(row=row, column=2, value=value)
        vc.font      = _font()
        vc.fill      = _fill("FFFFFF")
        vc.alignment = _align("left", wrap=True)
        vc.border    = _thin_border()
        blank = ws_cc.cell(row=row, column=3, value="")
        blank.border = _thin_border()

    _banner(1, "Collateral Configuration — Important Note")

    _row(3,  "What are the collateral columns?",
             "Columns E onward (Real Estate, Motor Vehicle …) represent the fair value of collateral "
             "pledged against each loan.  Each column name must be an exact, case-sensitive match to a "
             "Collateral Type name configured in your workspace (Workspace → Admin → Collateral Types).", True)
    _row(4,  "How to configure Collateral Types",
             "Navigate to Workspace → Admin → Collateral Types and create each category your portfolio uses.  "
             "The 'Name' field you enter there must match the column header in this file exactly.")
    _row(5,  "What is Haircut?",
             "A percentage reduction applied to the collateral fair value to account for liquidation costs, "
             "market discounts, and forced-sale risk.  E.g. a 20 % haircut on KES 1 000 000 gives "
             "KES 800 000 recoverable before discounting.")
    _row(6,  "What is Time to Realise?",
             "The expected number of months to liquidate this collateral type.  Longer realisation periods "
             "reduce present value because of the time-value-of-money discount applied.")
    _row(7,  "Discount formula",
             "For each collateral type:   Discounted Value = Fair Value × (1 − Haircut) × (1 + EIR) ^ (−Time_to_Realise)   "
             "where EIR is the loan's effective interest rate (column D).")
    _row(8,  "What if a collateral column is unrecognised?",
             "Validation error EC-02 blocks the run.  Either remove the column or add the collateral type "
             "in the admin screen so the system knows its haircut and time-to-realise values.")
    _row(9,  "What if a loan has no collateral of a particular type?",
             "Enter 0 (zero).  Blank cells are treated as 0 but may trigger a validation warning.  "
             "Do not delete the column — all configured collateral type columns must be present.")

    # Example table
    ws_cc.row_dimensions[11].height = 22
    ws_cc.cell(row=11, column=1, value="Example collateral types and typical IFRS 9 parameters").font = _font(bold=True)
    for ci, lbl in enumerate(["Collateral Type Name", "Typical Haircut %", "Typical Time to Realise (months)"], 1):
        c = ws_cc.cell(row=12, column=ci, value=lbl)
        c.font      = _font(bold=True, size=10, color=_HDR_FG)
        c.fill      = _fill(_HDR_BG)
        c.alignment = _align("center")
        c.border    = _header_border()
        ws_cc.row_dimensions[12].height = 22

    examples = [
        ("Real Estate",           "20 %",  "18"),
        ("Motor Vehicle",         "35 %",  "6"),
        ("Cash Deposit",          "0 %",   "1"),
        ("Corporate Guarantee",   "15 %",  "12"),
        ("Government Securities", "5 %",   "3"),
        ("Equipment / Machinery", "40 %",  "9"),
    ]
    for i, (name, hc, ttr) in enumerate(examples):
        bg = _ALT_ROW if i % 2 == 0 else "FFFFFF"
        for ci, val in enumerate([name, hc, ttr], 1):
            c = ws_cc.cell(row=13 + i, column=ci, value=val)
            c.fill      = _fill(bg)
            c.font      = _font()
            c.alignment = _align("center")
            c.border    = _thin_border()

    _make_instructions(
        wb,
        tab_color=_TAB_LGD,
        title="Loss Given Default (LGD) — Upload Guide",
        intro=(
            "Upload one LGD file per run (single workbook, 25 MB max).  Each row represents one active loan.  "
            "Collateral column names must exactly match the Collateral Types configured in your workspace.  "
            "See the 'Collateral Config' sheet in this workbook for detailed guidance."
        ),
        col_rows=[
            ("Customer ID",
             "Yes", "Text", "Any non-empty string",
             "A customer may have multiple loans.  Used for proportional collateral allocation: each loan's "
             "share of a customer's total collateral = Outstanding Amount / Total Customer Outstanding.",
             "CUST-001"),
            ("Loan ID",
             "Yes", "Text", "Unique per row",
             "Must be unique within this file.  Primary key for LGD results.",
             "LN-2025-0001"),
            ("Outstanding Amount",
             "Yes", "Number", "≥ 0",
             "Current outstanding principal balance.  Used to weight collateral allocation across "
             "loans belonging to the same customer.",
             "490000.00"),
            ("Effective Interest Rate (EIR)",
             "Yes", "Number", "0 – 1  (decimal fraction)",
             "Annual EIR expressed as a decimal, e.g. enter 0.14 for 14 %.  "
             "Used as the discount rate in the collateral present-value formula (EC-05: must be 0 – 1).",
             "0.14"),
            ("[Collateral Type Name]",
             "Yes*", "Number", "≥ 0",
             "*One column per configured Collateral Type.  Column name must be an exact match (EC-02).  "
             "Enter 0 if no collateral of this type is pledged for the loan.  "
             "Amount represents fair value before haircut.",
             "350000.00"),
        ],
        extra=[
            ("Collateral discount formula (applied per type per loan)", [
                "  Discounted Value = Fair Value  ×  (1 − Haircut)  ×  (1 + EIR) ^ (−Time_to_Realise_months)",
                "  where:  Haircut and Time_to_Realise come from Workspace → Admin → Collateral Types.",
                "          EIR comes from column D of this file.",
                "",
                "  The sum of discounted values across all collateral types = LGD recovery for that loan.",
                "  LGD % = 1 − (Recovery Amount / Outstanding Amount)  [capped at 0 % – 100 %].",
            ]),
        ],
    )


# ── EAD template ──────────────────────────────────────────────────────────────
def _build_ead(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "Data"
    ws.sheet_properties.tabColor = _TAB_EAD

    headers = [
        "Loan ID",
        "Customer ID",
        "SEGMENT",
        "Reporting Date",
        "Maturity Date",
        "Adjusted Maturity Date",
        "First Payment Date",
        "Outstanding Amount",
        "Repayment Frequency",
        "Staging",
        "Effective Interest Rate (EIR)",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)

    _DATE_COLS = {4, 5, 6, 7}

    rows = [
        # LID           CID        SEG          RepDate       MatDate       AdjMat        FPDate        OutAmt         Freq    Stage     EIR
        ("LN-2025-0001","CUST-001","Retail",    "2025-03-31","2029-03-31","2029-03-31","2022-04-30",  490_000.00, "MTH","Stage 1",0.1400),
        ("LN-2025-0002","CUST-001","Retail",    "2025-03-31","2028-06-30","2028-06-30","2021-07-31",  748_000.00, "MTH","Stage 2",0.1400),
        ("LN-2025-0003","CUST-002","Corporate", "2025-03-31","2030-12-31","2030-12-31","2023-01-31",3_100_000.00, "QTR","Stage 1",0.1200),
        ("LN-2025-0004","CUST-003","SME",       "2025-03-31","2026-09-30","2026-09-30","2023-10-31",1_060_000.00, "MTH","Stage 3",0.1600),
        ("LN-2025-0005","CUST-003","SME",       "2025-03-31","2027-03-31","2027-03-31","2022-04-30",  860_000.00, "MTH","Stage 1",0.1600),
        ("LN-2025-0006","CUST-004","Retail",    "2025-03-31","2031-03-31","2031-03-31","2024-04-30",1_250_000.00, "MTH","Stage 1",0.1300),
        ("LN-2025-0007","CUST-005","SME",       "2025-03-31","2026-03-31","2026-03-31","2023-04-30",  425_000.00, "QTR","Stage 2",0.1500),
        ("LN-2025-0008","CUST-005","SME",       "2025-03-31","2028-09-30","2028-09-30","2022-10-31",  980_000.00, "MTH","Stage 1",0.1500),
        ("LN-2025-0009","CUST-006","Corporate", "2025-03-31","2032-06-30","2032-06-30","2024-07-31",2_200_000.00, "QTR","Stage 1",0.1100),
        ("LN-2025-0010","CUST-007","SME",       "2025-03-31","2027-09-30","2027-09-30","2023-10-31",  630_000.00, "MTH","Stage 2",0.1700),
    ]

    for r, row_data in enumerate(rows, 2):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            if ci in _DATE_COLS:
                cell.number_format = _FMT_DATE
                cell.alignment     = _align("center")
            elif ci == 8:
                cell.number_format = _FMT_CURRENCY
                cell.alignment     = _align("right")
            elif ci == 11:
                cell.number_format = _FMT_DEC4
                cell.alignment     = _align("center")
            elif ci in (9, 10):
                cell.alignment = _align("center")
            else:
                cell.alignment = _align("center")

    _style_header_row(ws, 11)
    _style_data_rows(ws, 2, 11, 11)
    ws.freeze_panes = "A2"

    for w, col in zip([18, 14, 14, 18, 16, 24, 20, 22, 22, 14, 32], range(1, 12)):
        _col_width(ws, col, w)

    _add_dv(ws, _DV_FREQ,  "I2:I50000", "Repayment Frequency",
            "MTH = Monthly  |  QTR = Quarterly")
    _add_dv(ws, _DV_STAGE, "J2:J50000", "Staging",
            "Stage 1 = performing  |  Stage 2 = SICR  |  Stage 3 = default")

    _make_instructions(
        wb,
        tab_color=_TAB_EAD,
        title="Exposure at Default (EAD) — Upload Guide",
        intro=(
            "Upload one EAD file per run (single workbook, 25 MB max).  Each row is one active loan at "
            "the reporting date.  The engine performs a month-by-month forward balance walk (up to 299 months) "
            "computing the expected outstanding balance at each future period, then discounts the resulting "
            "ECL cash flows back to today."
        ),
        col_rows=[
            ("Loan ID",
             "Yes", "Text", "Unique, non-empty",
             "Must be unique within this file.  Primary key for the EAD rundown and for joining "
             "PD and LGD results at the loan level.",
             "LN-2025-0001"),
            ("Customer ID",
             "Yes", "Text", "Non-empty string",
             "Used to join LGD collateral data.  Must match the Customer ID in the LGD file.",
             "CUST-001"),
            ("SEGMENT",
             "Yes", "Text", "Must match a configured segment",
             "Case-sensitive match to a Segment name in Workspace → Admin → Segments.  "
             "Used to look up the segment-level PD curve.",
             "Retail"),
            ("Reporting Date",
             "Yes", "Date", "YYYY-MM-DD",
             "The as-at snapshot date for this loan.  All other dates are validated relative to this.",
             "2025-03-31"),
            ("Maturity Date",
             "Yes", "Date", "≥ Reporting Date",
             "Contractual maturity date.  Must be on or after Reporting Date (EC-03).  "
             "The rundown stops at this date.",
             "2029-03-31"),
            ("Adjusted Maturity Date",
             "Yes", "Date", "≥ First Payment Date",
             "Effective maturity after any restructuring or extensions.  "
             "Must be on or after First Payment Date.",
             "2029-03-31"),
            ("First Payment Date",
             "Yes", "Date", "≤ Reporting Date",
             "Date of the first scheduled repayment.  Must be on or before Reporting Date (EC-04).  "
             "Used to determine remaining loan term.",
             "2022-04-30"),
            ("Outstanding Amount",
             "Yes", "Number", "≥ 0",
             "Current outstanding principal balance at the reporting date.  "
             "Seed value for the forward balance walk.",
             "490000.00"),
            ("Repayment Frequency",
             "Yes", "Enum", "MTH  or  QTR",
             "Payment schedule frequency.  MTH = monthly, QTR = quarterly.  "
             "Other values trigger EC-09.  The balance walk always accrues monthly regardless.",
             "MTH"),
            ("Staging",
             "Yes", "Enum", "Stage 1  /  Stage 2  /  Stage 3",
             "IFRS 9 credit stage at the reporting date.  Determines which row of the segment PD "
             "transition matrix is used for lifetime marginal PD.",
             "Stage 1"),
            ("Effective Interest Rate (EIR)",
             "Yes", "Number", "0 – 1  (decimal fraction)",
             "Annual EIR as a decimal, e.g. 0.14 for 14 %.  Used for (a) instalment calculation "
             "and (b) ECL discounting (EC-05: must be in the range 0 – 1).",
             "0.14"),
        ],
        extra=[
            ("Cross-field date validation rules  (violations block the run)", [
                "EC-03   Maturity Date  ≥  Reporting Date",
                "        ✗  Reporting Date = 2025-03-31,  Maturity Date = 2024-12-31   ← maturity is in the past",
                "        ✓  Reporting Date = 2025-03-31,  Maturity Date = 2029-03-31",
                "",
                "EC-04   First Payment Date  ≤  Reporting Date",
                "        ✗  Reporting Date = 2025-03-31,  First Payment Date = 2025-06-30   ← first payment in the future",
                "        ✓  Reporting Date = 2025-03-31,  First Payment Date = 2022-04-30",
                "",
                "        Adjusted Maturity Date  ≥  First Payment Date",
                "        ✗  First Payment Date = 2022-04-30,  Adjusted Maturity Date = 2021-12-31",
                "        ✓  First Payment Date = 2022-04-30,  Adjusted Maturity Date = 2029-03-31",
            ]),
            ("Balance walk and instalment calculation", [
                "For MTH loans:  monthly instalment = PMT(EIR / 12, remaining_months, outstanding)",
                "For QTR loans:  quarterly instalment = PMT(EIR / 4,  remaining_quarters, outstanding)",
                "Monthly balance accrual always uses  EIR / 12  regardless of payment frequency.",
                "Stage 3 loans accrue missed-payment interest into the balance at each forward step.",
                "The rundown produces up to 299 monthly EAD snapshots per loan.",
            ]),
            ("Accepted aliases for column names", [
                "Staging column also accepted as:   Staging (Stage)",
                "EIR column also accepted as:       Effective Interest Rate   or   EIR",
            ]),
        ],
    )


# ── Entry point ───────────────────────────────────────────────────────────────
def main() -> None:
    _OUT_DIR.mkdir(parents=True, exist_ok=True)

    specs = [
        ("PD",  _build_pd,  "PD.xlsx"),
        ("LGD", _build_lgd, "LGD.xlsx"),
        ("EAD", _build_ead, "EAD.xlsx"),
    ]

    print("Generating ECL input templates …")
    for kind, builder, filename in specs:
        wb = Workbook()
        builder(wb)
        out_path = _OUT_DIR / filename
        wb.save(out_path)
        size_kb = out_path.stat().st_size // 1024
        print(f"  {kind:4s}  →  {out_path}  ({size_kb} KB)")

    print(f"\nDone.  Templates written to:  {_OUT_DIR}")


if __name__ == "__main__":
    main()
