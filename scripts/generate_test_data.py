#!/usr/bin/env python3
"""Generate a coherent dummy ECL portfolio (PD x2, LGD, EAD .xlsx files).

The five files produced under `ECL-Server/test_data/` are coordinated:
    * Loan IDs, Customer IDs, EIRs, segments and stages line up across
      PD (March snapshot), LGD and EAD.
    * Stage transitions across Jan/Feb/Mar/Apr 2025 are engineered so
      the IFRS 9 3x3 transition matrix is non-degenerate per segment,
      cure rate is populated, and a few loans drop off-books.

Deterministic: random.seed(42).

Run:
    python scripts/generate_test_data.py

Output: ECL-Server/test_data/*.xlsx
"""

from __future__ import annotations

import os
import random
import sys
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# -------------------------------------------------------------------------
# Configuration
# -------------------------------------------------------------------------

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "test_data"

REPORTING_DATE = date(2025, 3, 31)          # the LGD/EAD snapshot date
PD_MONTHS: list[date] = [                   # 4 monthly PD snapshots
    date(2025, 1, 31),
    date(2025, 2, 28),
    date(2025, 3, 31),
    date(2025, 4, 30),
]

SEGMENT_CONFIG = {
    # name      : (count, eir_range,           outstanding_range,        prefix)
    "Retail":    (30,    (0.13, 0.17),         (180_000, 1_200_000),     "RET"),
    "SME":       (20,    (0.13, 0.17),         (400_000, 2_500_000),     "SME"),
    "Corporate": (10,    (0.11, 0.14),         (1_500_000, 6_000_000),   "COR"),
}

COLLATERAL_TYPES = [
    "Real Estate",
    "Motor Vehicle",
    "Cash Deposit",
    "Corporate Guarantee",
]


# -------------------------------------------------------------------------
# Loan generation
# -------------------------------------------------------------------------

@dataclass
class Loan:
    loan_id: str
    customer_id: str
    segment: str
    outstanding: float           # March 31 balance, used by LGD + EAD
    eir: float                   # decimal fraction, e.g. 0.14
    repayment_frequency: str     # "MTH" or "QTR"
    first_payment_date: date
    maturity_date: date
    stages_by_month: dict[date, str]   # month -> "Stage 1"/"Stage 2"/"Stage 3"/None
    collateral: dict[str, float]       # per-type fair value (zeros allowed)


def _classify(idx: int, segment_size: int) -> str:
    """Assign one of the staging scenarios using deterministic buckets.

    The mix is tuned so the March-31 reporting cohort (used by LGD + EAD) has
    a healthy share of Stage 2 + Stage 3 loans (~25-30% of the portfolio),
    which makes the engine's ECL output visibly non-zero, while the four
    monthly PD snapshots still cover every transition type per segment.
    """
    pct = idx / segment_size
    if pct < 0.05:
        return "drop_off"           # S1, S1, S1, (absent in Apr)
    if pct < 0.15:
        return "deteriorate_slow"   # S1, S2, S2, S3   (Stage 2 in March)
    if pct < 0.25:
        return "deteriorate_fast"   # S1, S2, S3, S3   (Stage 3 in March)
    if pct < 0.32:
        return "cure"               # S2, S2, S1, S1
    if pct < 0.38:
        return "stage3_cure"        # S3, S2, S1, S1
    if pct < 0.48:
        return "stage3_persist"     # S3, S3, S3, S3   (Stage 3 in March)
    return "stable_s1"              # S1, S1, S1, S1   (~52% of portfolio)


_STAGE_TRACKS: dict[str, list[str | None]] = {
    "stable_s1":         ["Stage 1", "Stage 1", "Stage 1", "Stage 1"],
    "drop_off":          ["Stage 1", "Stage 1", "Stage 1", None],
    "deteriorate_slow":  ["Stage 1", "Stage 2", "Stage 2", "Stage 3"],
    "deteriorate_fast":  ["Stage 1", "Stage 2", "Stage 3", "Stage 3"],
    "cure":              ["Stage 2", "Stage 2", "Stage 1", "Stage 1"],
    "stage3_cure":       ["Stage 3", "Stage 2", "Stage 1", "Stage 1"],
    "stage3_persist":    ["Stage 3", "Stage 3", "Stage 3", "Stage 3"],
}


def _build_collateral(segment: str, outstanding: float, rng: random.Random) -> dict[str, float]:
    """Spread collateral across types with a segment-specific bias.

    Roughly 12% of loans get zero collateral (exercises the worst-case LGD path).
    """
    if rng.random() < 0.12:
        return {ct: 0.0 for ct in COLLATERAL_TYPES}

    # Total pledged value as a multiple of outstanding (under-, fully-, over-collateralised mix)
    coverage = rng.uniform(0.4, 1.4) * outstanding

    if segment == "Corporate":
        weights = {"Real Estate": 0.65, "Motor Vehicle": 0.0, "Cash Deposit": 0.05, "Corporate Guarantee": 0.30}
    elif segment == "SME":
        weights = {"Real Estate": 0.30, "Motor Vehicle": 0.15, "Cash Deposit": 0.25, "Corporate Guarantee": 0.30}
    else:  # Retail
        weights = {"Real Estate": 0.45, "Motor Vehicle": 0.45, "Cash Deposit": 0.10, "Corporate Guarantee": 0.0}

    # Apply jitter so two loans don't share an identical collateral profile.
    jittered = {ct: max(0.0, weights[ct] + rng.uniform(-0.05, 0.05)) for ct in COLLATERAL_TYPES}
    total_w = sum(jittered.values()) or 1.0
    return {ct: round(coverage * w / total_w, 2) for ct, w in jittered.items()}


def _month_offset(d: date, months: int) -> date:
    """Add (or subtract) whole months to a date, clamping the day to month-end."""
    y, m = d.year, d.month + months
    while m <= 0:
        y -= 1
        m += 12
    while m > 12:
        y += 1
        m -= 12
    # land on the last day of that month
    if m == 12:
        next_first = date(y + 1, 1, 1)
    else:
        next_first = date(y, m + 1, 1)
    last_day = (next_first - date(y, m, 1)).days
    return date(y, m, last_day)


def generate_loans() -> list[Loan]:
    rng = random.Random(42)
    loans: list[Loan] = []
    next_customer = 1
    customer_pool: list[str] = []

    for segment, (count, eir_range, outstanding_range, prefix) in SEGMENT_CONFIG.items():
        for i in range(count):
            loan_id = f"LN-2025-{prefix}-{i + 1:03d}"

            # Pool customers so ~25% of customers hold multiple loans (LGD pro-rata path).
            if rng.random() < 0.25 and customer_pool:
                customer_id = rng.choice(customer_pool)
            else:
                customer_id = f"CUST-{next_customer:04d}"
                customer_pool.append(customer_id)
                next_customer += 1

            scenario = _classify(i, count)
            track = _STAGE_TRACKS[scenario]
            stages_by_month = {m: track[idx] for idx, m in enumerate(PD_MONTHS)}

            outstanding = round(rng.uniform(*outstanding_range), 2)
            eir = round(rng.uniform(*eir_range), 4)

            # EAD dates: first payment 12-36 months in the past (EC-04 safe),
            # maturity 18-84 months in the future (EC-03 safe).
            first_payment_offset = -rng.randint(12, 36)
            maturity_offset = rng.randint(18, 84)
            first_payment_date = _month_offset(REPORTING_DATE, first_payment_offset)
            maturity_date = _month_offset(REPORTING_DATE, maturity_offset)

            # Corporate skews quarterly, the rest monthly.
            if segment == "Corporate":
                repayment_frequency = "QTR" if rng.random() < 0.6 else "MTH"
            else:
                repayment_frequency = "QTR" if rng.random() < 0.1 else "MTH"

            collateral = _build_collateral(segment, outstanding, rng)

            loans.append(
                Loan(
                    loan_id=loan_id,
                    customer_id=customer_id,
                    segment=segment,
                    outstanding=outstanding,
                    eir=eir,
                    repayment_frequency=repayment_frequency,
                    first_payment_date=first_payment_date,
                    maturity_date=maturity_date,
                    stages_by_month=stages_by_month,
                    collateral=collateral,
                )
            )
    return loans


# -------------------------------------------------------------------------
# Workbook writers
# -------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22


def write_pd_file(path: Path, loans: list[Loan], months: list[date]) -> int:
    """Write a PD workbook for the given months. Returns row count."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = ["Loan ID", "SEGMENT", "Reporting Month", "Staging", "Loan Amount"]
    ws.append(headers)

    rng = random.Random(hash(path.name) & 0xFFFF)  # per-file balance jitter, stable
    rows = 0
    for loan in loans:
        # Mild month-on-month amortisation so Loan Amount evolves realistically.
        base = loan.outstanding
        for m_idx, month in enumerate(months):
            stage = loan.stages_by_month.get(month)
            if stage is None:
                continue  # loan is off-books for this snapshot (drop_off scenario in Apr)
            # March-31 row keeps the exact LGD/EAD outstanding; other months drift slightly.
            if month == REPORTING_DATE:
                amt = round(loan.outstanding, 2)
            else:
                # ~1-2% noise around the March value (older months a touch higher, future a touch lower).
                month_distance = (month.year - REPORTING_DATE.year) * 12 + (month.month - REPORTING_DATE.month)
                drift = 1.0 - 0.012 * month_distance + rng.uniform(-0.005, 0.005)
                amt = round(base * drift, 2)
            ws.append([loan.loan_id, loan.segment, month, stage, amt])
            rows += 1

    _style_header(ws, len(headers))
    # Format the date column as YYYY-MM-DD for predictable validator parsing.
    for row in range(2, rows + 2):
        ws.cell(row=row, column=3).number_format = "yyyy-mm-dd"
    wb.save(path)
    return rows


def write_lgd_file(path: Path, loans: list[Loan]) -> int:
    """LGD uses the March 31 snapshot only — loans absent that month are excluded."""
    active = [l for l in loans if l.stages_by_month.get(REPORTING_DATE) is not None]

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = ["Customer ID", "Loan ID", "Outstanding Amount", "Effective Interest Rate (EIR)"]
    headers.extend(COLLATERAL_TYPES)
    ws.append(headers)

    for loan in active:
        row = [
            loan.customer_id,
            loan.loan_id,
            round(loan.outstanding, 2),
            loan.eir,
        ]
        row.extend(round(loan.collateral.get(ct, 0.0), 2) for ct in COLLATERAL_TYPES)
        ws.append(row)

    _style_header(ws, len(headers))
    wb.save(path)
    return len(active)


def write_ead_file(path: Path, loans: list[Loan]) -> int:
    active = [l for l in loans if l.stages_by_month.get(REPORTING_DATE) is not None]

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    headers = [
        "Loan ID",
        "Customer ID",
        "SEGMENT",
        "Reporting Date",
        "Maturity Date",
        "Adjusted Maturity Date",
        "First Payment Date",
        "Outstanding Amount",
        "Repayment Frequency",
        "Staging",
        "Effective Interest Rate (EIR)",
    ]
    ws.append(headers)

    for loan in active:
        # Adjusted Maturity == Maturity for this dataset (no restructures simulated).
        ws.append([
            loan.loan_id,
            loan.customer_id,
            loan.segment,
            REPORTING_DATE,
            loan.maturity_date,
            loan.maturity_date,
            loan.first_payment_date,
            round(loan.outstanding, 2),
            loan.repayment_frequency,
            loan.stages_by_month[REPORTING_DATE],
            loan.eir,
        ])

    _style_header(ws, len(headers))
    for col_idx in (4, 5, 6, 7):
        for row in range(2, len(active) + 2):
            ws.cell(row=row, column=col_idx).number_format = "yyyy-mm-dd"
    wb.save(path)
    return len(active)


# -------------------------------------------------------------------------
# Coverage report (sanity-check the engineered scenarios)
# -------------------------------------------------------------------------

def _print_coverage(loans: list[Loan]) -> None:
    print("\nStage-transition coverage per segment (Jan -> Feb -> Mar -> Apr):")
    for segment in SEGMENT_CONFIG:
        seg_loans = [l for l in loans if l.segment == segment]
        transitions: dict[tuple[str, str], int] = {}
        for loan in seg_loans:
            for prev_m, next_m in zip(PD_MONTHS, PD_MONTHS[1:]):
                a = loan.stages_by_month.get(prev_m)
                b = loan.stages_by_month.get(next_m, "Off-books")
                if a is None:
                    continue
                key = (a, b or "Off-books")
                transitions[key] = transitions.get(key, 0) + 1
        print(f"  {segment} ({len(seg_loans)} loans):")
        for (a, b), n in sorted(transitions.items()):
            print(f"    {a:<8} -> {b:<10} : {n}")


# -------------------------------------------------------------------------
# Entry point
# -------------------------------------------------------------------------

def main() -> int:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    loans = generate_loans()
    print(f"Generated {len(loans)} loans across {len(SEGMENT_CONFIG)} segments.")

    pd_file_1 = OUTPUT_DIR / "PD_2025_Jan_Feb.xlsx"
    pd_file_2 = OUTPUT_DIR / "PD_2025_Mar_Apr.xlsx"
    lgd_file = OUTPUT_DIR / "LGD_2025_03.xlsx"
    ead_file = OUTPUT_DIR / "EAD_2025_03.xlsx"

    rows_pd_1 = write_pd_file(pd_file_1, loans, PD_MONTHS[:2])
    rows_pd_2 = write_pd_file(pd_file_2, loans, PD_MONTHS[2:])
    rows_lgd = write_lgd_file(lgd_file, loans)
    rows_ead = write_ead_file(ead_file, loans)

    print(f"  PD  (Jan/Feb): {pd_file_1.name:<28} -> {rows_pd_1} rows")
    print(f"  PD  (Mar/Apr): {pd_file_2.name:<28} -> {rows_pd_2} rows")
    print(f"  LGD          : {lgd_file.name:<28} -> {rows_lgd} rows")
    print(f"  EAD          : {ead_file.name:<28} -> {rows_ead} rows")

    _print_coverage(loans)
    print(f"\nAll files written to: {OUTPUT_DIR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
